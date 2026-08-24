#!/usr/bin/env python3
"""RDB parsing and fixed-clock morphology metrics.

RDB labels are inclusive P/QRS/T regions.  They contain no labeled peaks.
This module resolves released floating coordinates to the 500 Hz sample grid
with deterministic half-up rounding and quarantines invalid intervals rather
than silently repairing them.
"""

from __future__ import annotations

import csv
import hashlib
import math
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import load_workbook


LEADS = ("i", "ii", "iii", "avr", "avl", "avf", "v1", "v2", "v3", "v4", "v5", "v6")
DISPLAY_LEADS = ("I", "II", "III", "aVR", "aVL", "aVF", "V1", "V2", "V3", "V4", "V5", "V6")
WAVES = {0: "P", 1: "QRS", 2: "T"}
OBSERVED = (0, 1, 7)
DERIVED_LIMB = (2, 3, 4, 5)
PRIMARY_PRECORDIAL = (6, 8, 9, 10, 11)
TARGET_FS = 500
TARGET_SAMPLES = 5000
SAMPLE_MS = 2.0
RAW_UNITS_PER_MV = 1000.0


def lead_role(index: int) -> str:
    if index in OBSERVED:
        return "observed_control"
    if index in DERIVED_LIMB:
        return "derived_limb_control"
    if index in PRIMARY_PRECORDIAL:
        return "primary_missing_precordial"
    raise ValueError(f"unclassified lead index {index}")


def round_half_up(value: float) -> int:
    """Resolve nonnegative annotation coordinates to the sampled signal grid."""
    if not math.isfinite(value) or value < 0:
        raise ValueError(f"invalid sample coordinate: {value}")
    return math.floor(value + 0.5)


def finite_pearson(reference: np.ndarray, reconstruction: np.ndarray) -> float:
    left = np.asarray(reference, dtype=np.float64)
    right = np.asarray(reconstruction, dtype=np.float64)
    left -= left.mean()
    right -= right.mean()
    denominator = math.sqrt(float(np.square(left).sum() * np.square(right).sum()))
    return float(np.dot(left, right) / denominator) if denominator > 1e-15 else float("nan")


def finite_quantile(values: Iterable[float], probability: float) -> float | None:
    array = np.asarray(list(values), dtype=np.float64)
    array = array[np.isfinite(array)]
    return float(np.quantile(array, probability)) if len(array) else None


def finite_mean(values: Iterable[float]) -> float | None:
    array = np.asarray(list(values), dtype=np.float64)
    array = array[np.isfinite(array)]
    return float(array.mean()) if len(array) else None


def finite_rmse(values: Iterable[float]) -> float | None:
    array = np.asarray(list(values), dtype=np.float64)
    array = array[np.isfinite(array)]
    return float(np.sqrt(np.mean(np.square(array)))) if len(array) else None


def sha256_file(path: Path, chunk_size: int = 8 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def wave_features(values: np.ndarray, onset: int, offset: int) -> dict[str, float]:
    """Features on the exact inclusive RDB interval; no peak is invented."""
    if not 0 <= onset < offset < len(values):
        raise ValueError("invalid wave interval")
    window = np.asarray(values[onset : offset + 1], dtype=np.float64)
    baseline = np.linspace(window[0], window[-1], len(window), dtype=np.float64)
    residual = window - baseline
    return {
        "signed_area_mv_ms": float(np.trapezoid(residual, dx=SAMPLE_MS)),
        "absolute_area_mv_ms": float(np.trapezoid(np.abs(residual), dx=SAMPLE_MS)),
    }


def st_features(values: np.ndarray, qrs_end: int, t_onset: int) -> dict[str, float | None]:
    """QRS offset is END; the J/ST-start sample is END+1."""
    j_sample = qrs_end + 1
    if not 0 <= qrs_end < j_sample < t_onset < len(values):
        raise ValueError("invalid ST interval")
    window = np.asarray(values[j_sample:t_onset], dtype=np.float64)
    result: dict[str, float | None] = {
        "qrs_offset_mv": float(values[qrs_end]),
        "j_mv": float(values[j_sample]),
        "mean_mv": float(window.mean()),
        "area_mv_ms": float(np.trapezoid(window, dx=SAMPLE_MS)),
    }
    for delay in (20, 40, 60, 80):
        sample = j_sample + round(delay / SAMPLE_MS)
        result[f"j{delay}_mv"] = float(values[sample]) if sample < t_onset else None
    return result


def read_mapping(path: Path) -> list[dict[str, str | None]]:
    sheet = load_workbook(path, read_only=True, data_only=True)["mapping"]
    rows = list(sheet.iter_rows(values_only=True))
    expected = ("rdb_record_id", "chapman_record_id", "rdb_rhythm", "chapman_rhythm", "note")
    if tuple(rows[0]) != expected:
        raise ValueError(f"unexpected mapping header: {rows[0]}")
    result = [dict(zip(expected, row, strict=True)) for row in rows[1:]]
    if len(result) != 2399 or len({r["rdb_record_id"] for r in result}) != 2399:
        raise ValueError("mapping must contain 2,399 unique RDB IDs")
    return result


def read_annotation(path: Path, audit: Counter[str], scope: str = "lead") -> np.ndarray:
    rows: list[tuple[int, int, int]] = []
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = next(reader, None)
        if not header or [part.strip().upper() for part in header] != ["TYPE", "START", "END"]:
            raise ValueError(f"missing RDB annotation header: {path}")
        audit["headers"] += 1
        audit[f"{scope}_headers"] += 1
        for line_number, row in enumerate(reader, 2):
            audit["annotation_rows"] += 1
            audit[f"{scope}_annotation_rows"] += 1
            if len(row) != 3:
                audit["excluded_wrong_columns"] += 1
                audit[f"{scope}_excluded_wrong_columns"] += 1
                continue
            try:
                kind = int(row[0]); raw_onset = float(row[1]); raw_offset = float(row[2])
            except ValueError:
                audit["excluded_parse_error"] += 1
                audit[f"{scope}_excluded_parse_error"] += 1
                continue
            if kind not in WAVES or not math.isfinite(raw_onset) or not math.isfinite(raw_offset):
                audit["excluded_invalid_value"] += 1
                audit[f"{scope}_excluded_invalid_value"] += 1
                continue
            if raw_onset < 0 or raw_offset < 0:
                audit["excluded_out_of_bounds"] += 1
                audit[f"{scope}_excluded_out_of_bounds"] += 1
                continue
            onset, offset = round_half_up(raw_onset), round_half_up(raw_offset)
            adjustment = max(abs(onset - raw_onset), abs(offset - raw_offset))
            audit["fractional_rows"] += int(raw_onset != int(raw_onset) or raw_offset != int(raw_offset))
            audit[f"{scope}_fractional_rows"] += int(raw_onset != int(raw_onset) or raw_offset != int(raw_offset))
            audit["max_rounding_adjustment_microsamples"] = max(
                audit["max_rounding_adjustment_microsamples"], round(adjustment * 1_000_000)
            )
            if raw_offset < raw_onset or offset < onset:
                audit["excluded_reversed"] += 1
                audit[f"{scope}_excluded_reversed"] += 1
                continue
            if offset == onset:
                audit["excluded_zero_length"] += 1
                audit[f"{scope}_excluded_zero_length"] += 1
                continue
            if onset >= TARGET_SAMPLES or offset >= TARGET_SAMPLES:
                audit["excluded_out_of_bounds"] += 1
                audit[f"{scope}_excluded_out_of_bounds"] += 1
                continue
            rows.append((kind, onset, offset))
            audit[f"valid_{WAVES[kind].lower()}_intervals"] += 1
            audit[f"{scope}_valid_{WAVES[kind].lower()}_intervals"] += 1
    return np.asarray(rows, dtype=np.int32).reshape(-1, 3)


def dataset_identity(root: Path, mapping_path: Path, selected: list[dict[str, str | None]]) -> str:
    digest = hashlib.sha256()
    for path in (root / "README.md", mapping_path):
        digest.update(path.name.encode()); digest.update(sha256_file(path).encode())
    for row in selected:
        record_id = str(row["rdb_record_id"])
        for path in [root / "dat_csv" / f"{record_id}.csv", *(
            root / "ann_txt" / f"{record_id}.{lead}.txt" for lead in (*LEADS, "all")
        )]:
            if not path.is_file():
                raise FileNotFoundError(path)
            digest.update(path.name.encode()); digest.update(sha256_file(path).encode())
    return digest.hexdigest()


def load_rdb(
    root: Path,
    mapping_path: Path,
    max_records: int = 0,
    include_flagged_duplicates: bool = False,
) -> tuple[list[dict[str, Any]], str, dict[str, int]]:
    mapping = read_mapping(mapping_path)
    selected = [r for r in mapping if include_flagged_duplicates or r["note"] != "duplicate_rdb_record"]
    excluded_flagged_duplicates = len(mapping) - len(selected)
    if max_records:
        selected = selected[:max_records]
    audit: Counter[str] = Counter()
    audit["mapping_rows"] = len(mapping)
    audit["mapping_unique_chapman_records"] = len({r["chapman_record_id"] for r in mapping})
    audit["excluded_mapping_flagged_duplicates"] = excluded_flagged_duplicates
    records: list[dict[str, Any]] = []
    for row in selected:
        record_id = str(row["rdb_record_id"])
        signal = np.loadtxt(root / "dat_csv" / f"{record_id}.csv", delimiter=",", dtype=np.float32)
        if signal.shape != (TARGET_SAMPLES, 12) or not np.isfinite(signal).all():
            raise ValueError(f"invalid RDB signal: {record_id} {signal.shape}")
        signal = (signal.T / RAW_UNITS_PER_MV).astype(np.float32)
        streams = [read_annotation(root / "ann_txt" / f"{record_id}.{lead}.txt", audit, "lead") for lead in LEADS]
        # Consensus is audited but never substituted for lead-specific labels.
        consensus = read_annotation(root / "ann_txt" / f"{record_id}.all.txt", audit, "consensus")
        audit["lead_streams_equal_consensus"] += sum(np.array_equal(stream, consensus) for stream in streams)
        audit["records"] += 1
        audit["lead_streams"] += 12
        records.append({
            "record_id": record_id,
            "chapman_record_id": str(row["chapman_record_id"]),
            "released_rhythm": str(row["rdb_rhythm"]),
            "canonical_rhythm": str(row["chapman_rhythm"]),
            "signal": signal,
            "annotations": streams,
        })
    if not records:
        raise RuntimeError("no RDB records loaded")
    audit["raw_signal_units_per_mv"] = int(RAW_UNITS_PER_MV)
    return records, dataset_identity(root, mapping_path, selected), dict(audit)


def interval_metric(reference: np.ndarray, predicted: np.ndarray, onset: int, offset: int) -> dict[str, float]:
    left = np.asarray(reference[onset : offset + 1], dtype=np.float64)
    right = np.asarray(predicted[onset : offset + 1], dtype=np.float64)
    difference = right - left
    ref_features = wave_features(reference, onset, offset)
    pred_features = wave_features(predicted, onset, offset)
    return {
        "onset_error_mv": float(predicted[onset] - reference[onset]),
        "offset_error_mv": float(predicted[offset] - reference[offset]),
        "window_pearson": finite_pearson(left, right),
        "window_rmse_mv": float(np.sqrt(np.mean(np.square(difference)))),
        "window_mae_mv": float(np.mean(np.abs(difference))),
        "window_max_abs_error_mv": float(np.max(np.abs(difference))),
        "signed_area_error_mv_ms": pred_features["signed_area_mv_ms"] - ref_features["signed_area_mv_ms"],
        "absolute_area_error_mv_ms": pred_features["absolute_area_mv_ms"] - ref_features["absolute_area_mv_ms"],
    }


def next_t_onset(intervals: np.ndarray, qrs_end: int) -> int | None:
    candidates = [int(start) for kind, start, _ in intervals if kind == 2 and start > qrs_end + 1]
    return min(candidates) if candidates else None
